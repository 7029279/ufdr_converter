import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


WORKSPACE_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = WORKSPACE_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

import ufdr_converter as converter


class UfdrConverterTests(unittest.TestCase):
    """展開済みテンポラリを使ってUFDR変換処理を検証する。"""

    @classmethod
    def setUpClass(cls):
        """テストで共通利用するサンプルUFDRの場所を固定する。"""
        cls.sample_ufdr = WORKSPACE_ROOT / "data" / "in" / "iphone.ufdr"
        if not cls.sample_ufdr.exists():
            raise FileNotFoundError(cls.sample_ufdr)

    def setUp(self):
        """各テストごとに独立した作業ディレクトリを用意する。"""
        self._temporary_directory = tempfile.TemporaryDirectory()
        self.addCleanup(self._temporary_directory.cleanup)
        self.temp_root = Path(self._temporary_directory.name)

    def extract_sample_to_temp(self):
        """サンプルUFDRをテンポラリへ展開してルートXMLまで返す。"""
        extract_dir = self.temp_root / "unzipped"
        converter.extract_ufdr(self.sample_ufdr, extract_dir)
        report_root = converter.parse_report(extract_dir / "report.xml")
        return extract_dir, report_root

    def test_extract_ufdr_creates_report_and_files(self):
        """UFDR展開でreport.xmlとファイル群が生成されることを確認する。"""
        extract_dir, _report_root = self.extract_sample_to_temp()

        self.assertTrue((extract_dir / "report.xml").exists())
        self.assertTrue((extract_dir / "files").exists())

    def test_build_tagged_file_index_reads_archive_paths(self):
        """taggedFilesからローカルパス付きの索引が作られることを確認する。"""
        _extract_dir, report_root = self.extract_sample_to_temp()

        tagged_files = converter.build_tagged_file_index(report_root)

        self.assertEqual(2, len(tagged_files))
        self.assertIn("b231534c-b43b-477d-863e-412342341", tagged_files)
        self.assertEqual(
            "files/Image/4454825783_dbcb233af5_b.jpg",
            tagged_files["b231534c-b43b-477d-863e-412342341"]["ArchivePath"],
        )

    def test_extract_chat_exports_writes_media_metadata(self):
        """チャット添付画像がアプリ別フォルダとメタデータへ出力されることを確認する。"""
        extract_dir, report_root = self.extract_sample_to_temp()
        output_dir = self.temp_root / "out"
        top_level_models = converter.collect_top_level_models(report_root)
        tagged_files = converter.build_tagged_file_index(report_root)

        attachment_ids = converter.extract_chat_exports(top_level_models["Chat"], tagged_files, extract_dir, output_dir)

        metadata_path = output_dir / "SNS_Chat_Media" / "iMessage" / "Metadata.xlsx"
        copied_media = output_dir / "SNS_Chat_Media" / "iMessage" / "4454825783_dbcb233af5_b.jpg"
        self.assertIn("b231534c-b43b-477d-863e-412342341", attachment_ids)
        self.assertTrue(metadata_path.exists())
        self.assertTrue(copied_media.exists())

        workbook = load_workbook(metadata_path, read_only=True)
        try:
            worksheet = workbook.active
            self.assertGreaterEqual(worksheet.max_row, 2)
        finally:
            workbook.close()

    def test_process_ufdr_creates_output_and_persistent_extract_dir(self):
        """本処理で成果物と同名の展開済みディレクトリが残ることを確認する。"""
        output_root = self.temp_root / "out"
        temp_root = self.temp_root / "temp"

        converter.process_ufdr(self.sample_ufdr, output_root, temp_root, "0001")

        self.assertTrue((output_root / "0001" / "Chat_Messages.xlsx").exists())
        self.assertTrue((output_root / "0001" / "Contacts.xlsx").exists())
        self.assertTrue((temp_root / "0001" / "report.xml").exists())
        self.assertTrue((temp_root / "0001" / "files").exists())

    def test_format_case_name_zero_pads_numbers(self):
        """ケース名が4桁ゼロ埋めになることを確認する。"""
        self.assertEqual("0001", converter.format_case_name(1))
        self.assertEqual("0042", converter.format_case_name(42))


if __name__ == "__main__":
    unittest.main()